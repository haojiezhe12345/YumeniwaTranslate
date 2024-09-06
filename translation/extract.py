import os
import re
import json
import openpyxl
import openpyxl.styles
import csv

srcDir = 'data/scenario'

db = {}
charas = []
lastChara = ''
skip = False

for fn in os.listdir(srcDir):
    db[fn] = {}
    with open(os.path.join(srcDir, fn), encoding='utf-8') as f:
        for line in f.readlines():
            if '[iscript]' in line:
                skip = True
            if '[endscript]' in line:
                skip = False
            if skip:
                continue

            txt = ''
            if line[0] == '#':
                lastChara = line[1:-1]
                if lastChara and not lastChara in charas:
                    charas.append(lastChara)
            elif not line[0] in ['\n', '@', '[', '*', ';', '\t', ' ', '﻿']:
                line = line.replace('[l][r]', '\n').replace('[l]', '\n').replace('[r]', '\n').replace('[lr]', '\n').replace('[p]', '\n').strip()
                if line:
                    txt = line
            else:
                r = [
                    r'\[.*text="(.*?)".*\]',
                    r'\[.*title name="(.*?)".*\]',
                    r'\[link.*\](.*)\[endlink\]',
                    r'\[font.*\](.*)',
                ]
                for r1 in r:
                    m = re.findall(r1, line)
                    if m:
                        txt = m[0]
            if txt:
                txts = txt.split('\n')
                for txt1 in txts:
                    txt1 = txt1.strip()
                    db[fn][txt1] = {
                        'chara': lastChara,
                        'translation': '',
                    }
    if not db[fn]:
        del db[fn]


with open('part_translated.csv', encoding='utf-8') as f:
    reader = csv.reader(f, delimiter='\t')
    workingfile = ''
    for row in reader:
        if row[0] == 'ScenarioName':
            workingfile = row[1]
        else:
            srctxt = row[0].replace('&nbsp;', ' ').strip()
            if workingfile in db and srctxt in db[workingfile]:
                db[workingfile][srctxt]['translation'] = row[1]
            elif srctxt:
                print(f'Failed matching from csv: "{srctxt}"')


with open(f'extracted.json', 'w', encoding='utf-8') as f:
    json.dump(db, f, ensure_ascii=False, indent=2)

with open(f'extracted_chara.txt', 'w', encoding='utf-8') as f:
    for chara in charas:
        f.write(chara + '\n')


wb = openpyxl.Workbook()
index = wb.active
index.title = 'Index'
index.column_dimensions['A'].width = 30

for i, fn in enumerate(db):
    ws = wb.create_sheet(fn)
    index.cell(row=i+1, column=1, value=fn).hyperlink = f'#{fn}!A1'
    index.cell(row=i+1, column=1).style = 'Hyperlink'
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 36
    ws.cell(row=1, column=1, value='←返回目录').hyperlink = '#Index!A1'
    ws.cell(row=1, column=1).style = 'Hyperlink'
    ws.cell(row=1, column=2, value='原文').alignment = openpyxl.styles.Alignment(horizontal="center")
    ws.cell(row=1, column=3, value='译文').alignment = openpyxl.styles.Alignment(horizontal="center")

    for j, txt in enumerate(db[fn]):
        ws.cell(row=j+2, column=1, value=db[fn][txt]['chara'])
        ws.cell(row=j+2, column=2, value=txt).alignment = openpyxl.styles.Alignment(wrap_text=True)
        ws.cell(row=j+2, column=3, value=db[fn][txt]['translation']).alignment = openpyxl.styles.Alignment(wrap_text=True)

wb.save('extracted.xlsx')
