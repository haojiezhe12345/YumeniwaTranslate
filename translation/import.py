import os
import json
import openpyxl

srcDir = 'data/scenario'
outDir = '../data/scenario'

db = {}
wb = openpyxl.load_workbook(R"F:\Downloads\ゆめのおにわでchasing - 文本汉化表.xlsx")

for sheet in wb.worksheets:
    if sheet.title != 'Index':
        db[sheet.title] = {}
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=3):
            if row[0].value:
                src = str(row[0].value)
                tld = str(row[1].value or '')
                db[sheet.title][src] = tld
        totalNum = len(db[sheet.title])
        translatedNum = len([1 for txt in db[sheet.title] if db[sheet.title][txt]])
        print(f'{sheet.title:<30} {translatedNum:>3} / {totalNum:<3}   {translatedNum / totalNum * 100:>3.0f}%')
        db[sheet.title] = dict(sorted(list(db[sheet.title].items()), key=lambda a: len(a[0]), reverse=True))

        with open(os.path.join(srcDir, sheet.title), encoding='utf-8') as f:
            ks0 = f.read()
            ks = ks0
            for src in db[sheet.title]:
                tld = db[sheet.title][src]
                if not src in ks:
                    raise Exception(src)
                if tld:
                    ks = ks.replace(src, tld)

            for name, nameTL in [
                ('【まどか】', '【圆】'),
                ('【ほむら】', '【焰】'),
                ('【さやか】', '【沙耶加】'),
                ('【杏子】', ''),
                ('【エイミー】', '【艾米】'),
                ('【？？？】', ''),
                ('【まどか&ほむら】', '【圆&焰】'),
                ('【詢子】', '【询子】'),
                ('【マミ】', '【麻美】'),
                ('【なぎさ】', '【渚】'),
                ('【？？】', ''),
                ('【マミ&なぎさ】', '【麻美&渚】'),
                # ('【ほむら】　', '【焰】　'),
                ('【ほむら&マミ&なぎさ】', '【焰&麻美&渚】'),
                ('【仁美】', ''),
                ('【先生】', '【老师】'),
                ('【女子Ａ】', ''),
                ('【男子Ｂ】', ''),
                ('【一同】', ''),
                ('【中沢】', '【中泽】'),
                ('【女子Ｂ】', ''),
                ('【男子Ａ】', ''),
                ('【マドカ】', '【圆】'),
                ('【まどか&ほむら&マミ】', '【圆&焰&麻美】'),
            ]:
                if nameTL:
                    ks = ks.replace(name, nameTL)

        if ks != ks0:
            with open(os.path.join(outDir, sheet.title), 'w', encoding='utf-8') as f:
                f.write(ks)


with open(f'translated.json', 'w', encoding='utf-8') as f:
    json.dump(db, f, ensure_ascii=False, indent=2)
